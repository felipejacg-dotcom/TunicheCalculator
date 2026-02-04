import tkinter as tk
from tkinter import messagebox, filedialog
from PIL import Image, ImageTk, ImageFilter
import os
import sys
from datetime import datetime

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# REGISTROS PARA INFORME EXCEL
# -----------------------------
registros = []  # aquí se acumulan todas las validaciones


# -----------------------------
# RUTA BASE (portable para EXE)
# -----------------------------
def base_dir():
    # si está empaquetado con PyInstaller
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    # si corre como .py normal
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = base_dir()

# Archivos (deja estos archivos en la misma carpeta del .py / .exe)
RUTA_FONDO = os.path.join(BASE_DIR, ""fondo.jpg"")
RUTA_ICONO_ICO = os.path.join(BASE_DIR, "icon.ico")


# --- 1. FUNCIÓN DE CÁLCULO ---
def calcular():
    try:
        alto = float(entry_alto.get())
        peso_envase = float(entry_peso_envase.get())
        peso_fruta = float(entry_peso_fruta.get())
        bct = float(entry_bct.get())

        # Factor de seguridad editable (libre: 1,2,3...8,10, etc.)
        FACTOR = float(entry_factor.get())
        if FACTOR <= 0:
            raise ValueError("Factor inválido")

        ALTURA_MAX = 2400
        PALLET_BASE = 150

        peso_caja_total = peso_fruta + (peso_envase / 1000)
        altura_util = ALTURA_MAX - PALLET_BASE
        pisos = int(altura_util / alto)

        # Validación mínima para evitar pisos inválidos
        if pisos < 2:
            etiqueta_resultado.config(text="⚠️ DATOS NO VÁLIDOS", bg="#f39c12", fg="white")
            ventana.update()
            messagebox.showwarning(
                "Advertencia",
                "Con ese alto de caja, no se puede apilar de forma válida (pisos < 2). Revisa el alto."
            )
            return

        cajas_encima = pisos - 1
        peso_sobre_base = cajas_encima * peso_caja_total
        resistencia_req = peso_sobre_base * FACTOR
        margen = bct - resistencia_req

        resultado_texto = f"Pisos en Pallet: {pisos}\n"
        resultado_texto += f"Factor Seguridad: {FACTOR:g}\n"
        resultado_texto += f"Resistencia Necesaria: {resistencia_req:.1f} kg\n"
        resultado_texto += f"Tu Caja: {bct:.1f} kg\n\n"

        # SEMÁFORO VISUAL
        if bct >= resistencia_req:
            estado = "APROBADA"
            etiqueta_resultado.config(text="✅ CAJA APROBADA", bg="#28a745", fg="white")
            ventana.update()
            messagebox.showinfo(
                "Resultado Exitoso",
                resultado_texto + f"APROBADA.\nTe sobran {margen:.1f} kg."
            )
        else:
            estado = "RECHAZADA"
            etiqueta_resultado.config(text="❌ CAJA RECHAZADA", bg="#dc3545", fg="white")
            ventana.update()
            messagebox.showwarning(
                "Alerta de Calidad",
                resultado_texto + f"RECHAZADA.\nTe faltan {abs(margen):.1f} kg."
            )

        # Guardar registro para el Excel (incluye el factor)
        registros.append({
            "FechaHora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "AltoCaja_mm": alto,
            "PesoEnvase_gr": peso_envase,
            "PesoFruta_kg": peso_fruta,
            "BCT_kg": bct,
            "FactorSeguridad": FACTOR,
            "Pisos": pisos,
            "ResistenciaNecesaria_kg": round(resistencia_req, 1),
            "Margen_kg": round(margen, 1),
            "Estado": estado
        })

        lbl_contador.config(text=f"Registros guardados: {len(registros)}")

    except ValueError:
        messagebox.showerror("Error", "Por favor ingresa todos los números correctamente (incluye Factor > 0).")
    except Exception as e:
        messagebox.showerror("Error", f"Detalle: {e}")


# --- 2. FUNCIÓN PARA GENERAR EXCEL ---
def generar_informe_excel():
    try:
        if not registros:
            messagebox.showwarning("Sin datos", "Aún no hay validaciones guardadas. Primero valida al menos 1 caja.")
            return

        nombre_sugerido = f"Informe_Validacion_Cajas_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel (*.xlsx)", "*.xlsx")],
            initialfile=nombre_sugerido,
            title="Guardar informe Excel"
        )

        if not ruta:
            return  # usuario canceló

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe"

        headers = [
            "Fecha/Hora", "Alto Caja (mm)", "Peso Envase (gr)", "Peso Fruta (kg)",
            "BCT (kg)", "Factor Seguridad", "Pisos",
            "Resistencia Necesaria (kg)", "Margen (kg)", "Estado"
        ]
        ws.append(headers)

        # Estilos
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")

        for col, _h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Datos
        for r in registros:
            ws.append([
                r["FechaHora"],
                r["AltoCaja_mm"],
                r["PesoEnvase_gr"],
                r["PesoFruta_kg"],
                r["BCT_kg"],
                r["FactorSeguridad"],
                r["Pisos"],
                r["ResistenciaNecesaria_kg"],
                r["Margen_kg"],
                r["Estado"]
            ])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Ajustar ancho de columnas + alineación
        for col in range(1, len(headers) + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

        wb.save(ruta)
        messagebox.showinfo("Listo", f"Informe Excel generado correctamente:\n{ruta}")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el Excel.\nDetalle: {e}")


# --- 3. CONFIGURACIÓN ---
ANCHO = 950
ALTO = 700

FONT_TITULO = ("Segoe UI", 20, "bold")
FONT_TEXTO = ("Segoe UI", 11)
FONT_BOTON = ("Segoe UI", 12, "bold")


ventana = tk.Tk()
ventana.title("Sistema de Validación Tuniche")
ventana.geometry(f"{ANCHO}x{ALTO}")
ventana.resizable(False, False)

# --- ICONO (.ico) ---
try:
    if os.path.exists(RUTA_ICONO_ICO):
        ventana.iconbitmap(RUTA_ICONO_ICO)
except Exception as e:
    print("No se pudo cargar icon.ico:", e)


# --- 4. FONDO ---
try:
    if not os.path.exists(RUTA_FONDO):
        messagebox.showwarning(
            "Fondo no encontrado",
            f"No se encontró la imagen de fondo:\n{RUTA_FONDO}\n\n"
            "Deja 'fondo.jpg' en la misma carpeta del .exe/.py."
        )
        ventana.configure(bg="#f0f0f0")
    else:
        img_original = Image.open(RUTA_FONDO)

        ratio_ventana = ANCHO / ALTO
        ratio_imagen = img_original.width / img_original.height

        if ratio_imagen > ratio_ventana:
            nuevo_alto = ALTO
            nuevo_ancho = int(nuevo_alto * ratio_imagen)
        else:
            nuevo_ancho = ANCHO
            nuevo_alto = int(nuevo_ancho / ratio_imagen)

        img_redimensionada = img_original.resize((nuevo_ancho, nuevo_alto), Image.Resampling.LANCZOS)
        left = (nuevo_ancho - ANCHO) / 2
        top = (nuevo_alto - ALTO) / 2
        img_crop = img_redimensionada.crop((left, top, left + ANCHO, top + ALTO))
        img_nublada = img_crop.filter(ImageFilter.GaussianBlur(radius=4))

        fondo_tk = ImageTk.PhotoImage(img_nublada, master=ventana)
        etiqueta_fondo = tk.Label(ventana, image=fondo_tk)
        etiqueta_fondo.image = fondo_tk  # IMPORTANTÍSIMO: evita que se borre
        etiqueta_fondo.place(x=0, y=0, relwidth=1, relheight=1)

except Exception as e:
    messagebox.showerror("Error cargando fondo", str(e))
    ventana.configure(bg="#f0f0f0")



# --- 5. PANEL CENTRAL ---
panel = tk.Frame(ventana, bg="#ffffff", bd=1, relief="solid")
panel.place(relx=0.5, rely=0.5, anchor="center", width=700, height=620)

tk.Label(panel, text="VALIDACIÓN DE CAJAS", font=FONT_TITULO, bg="#ffffff", fg="#2c3e50").pack(pady=(25, 10))
tk.Label(panel, text="DATOS TÉCNICOS", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#95a5a6").pack(pady=(0, 10))


def crear_casilla(padre, texto_etiqueta):
    f = tk.Frame(padre, bg="white")
    f.pack(pady=5, padx=60, fill="x")
    lbl = tk.Label(f, text=texto_etiqueta, font=FONT_TEXTO, bg="white", width=22, anchor="w", fg="#34495e")
    lbl.pack(side="left")
    ent = tk.Entry(f, font=FONT_TEXTO, bd=1, relief="solid")
    ent.config(highlightbackground="#bdc3c7", highlightcolor="#3498db")
    ent.pack(side="left", fill="x", expand=True, ipady=3)
    return ent


entry_alto = crear_casilla(panel, "Alto de Caja (mm):")
entry_peso_envase = crear_casilla(panel, "Peso Envase (gr):")
entry_peso_fruta = crear_casilla(panel, "Peso Fruta (kg):")
entry_bct = crear_casilla(panel, "Resistencia BCT (kg):")
entry_factor = crear_casilla(panel, "Factor Seguridad:")
entry_factor.insert(0, "8")  # valor por defecto

tk.Label(panel, text="PARÁMETROS", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#95a5a6").pack(pady=(18, 5))

lbl_config = tk.Label(
    panel,
    text="Pallet 2.40m  |  Factor Seguridad: 8",
    font=("Segoe UI", 12),
    bg="#ecf0f1",
    fg="#2c3e50",
    padx=15,
    pady=6
)
lbl_config.pack(pady=5, fill="x", padx=60)


def actualizar_factor_label(*_):
    try:
        f = float(entry_factor.get())
        lbl_config.config(text=f"Pallet 2.40m  |  Factor Seguridad: {f:g}")
    except:
        lbl_config.config(text="Pallet 2.40m  |  Factor Seguridad: -")


entry_factor.bind("<KeyRelease>", actualizar_factor_label)
actualizar_factor_label()

btn_validar = tk.Button(
    panel, text="VALIDAR CAJA", command=calcular,
    bg="#2980b9", fg="white", font=FONT_BOTON, bd=0,
    padx=20, pady=10, cursor="hand2"
)
btn_validar.pack(pady=(18, 10), fill="x", padx=60)

btn_excel = tk.Button(
    panel, text="GENERAR INFORME EXCEL", command=generar_informe_excel,
    bg="#27ae60", fg="white", font=FONT_BOTON, bd=0,
    padx=20, pady=10, cursor="hand2"
)
btn_excel.pack(pady=(0, 12), fill="x", padx=60)

# Contador de registros
lbl_contador = tk.Label(panel, text="Registros guardados: 0", font=("Segoe UI", 11, "bold"),
                        bg="#ffffff", fg="#2c3e50")
lbl_contador.pack(pady=(0, 8))

# ETIQUETA DE RESULTADO
etiqueta_resultado = tk.Label(panel, text="", font=("Segoe UI", 16, "bold"), bg="#ffffff", width=45)
etiqueta_resultado.pack(pady=(0, 18), ipady=10)

ventana.mainloop()


